import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from convertir_excel_a_pdf import convertir_excel_a_pdf
from abrir_carpeta_pdf import abrir_carpeta_pdf
from barra_de_carga import BarraDeCarga
from reemplazar_espacios_en_archivos import (
    reemplazar_espacios_en_archivos,
)


# Variable global para almacenar la ruta del archivo Excel seleccionado
archivo_excel_global = None
# -------------------------------------------------------------------------------


# Función para abrir un archivo Excel y actualizar las listas desplegables
def abrir_archivo_excel():
    global archivo_excel_global
    barra_progreso_archivos.reset()
    archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if archivo_excel:
        archivo_excel_global = archivo_excel
        workbook = openpyxl.load_workbook(archivo_excel, data_only=True)
        hojas = workbook.sheetnames
        lista_hojas_inicio["values"] = hojas
        lista_hojas_fin["values"] = hojas
        workbook.close()


# -------------------------------------------------------------------------------


# Función para convertir un rango específico de hojas de Excel a PDF
def convertir_rango_especifico(app, hoja_inicio, hoja_fin, barra_progreso_rango):
    global archivo_excel_global
    if archivo_excel_global:
        workbook = openpyxl.load_workbook(archivo_excel_global, data_only=True)
        todas_las_hojas = workbook.sheetnames
        try:
            inicio = todas_las_hojas.index(hoja_inicio)
            fin = todas_las_hojas.index(hoja_fin) + 1
            hojas_a_convertir = todas_las_hojas[inicio:fin]
        except ValueError:
            messagebox.showerror("Error", "Las hojas seleccionadas no son válidas.")
            return
        convertir_excel_a_pdf(
            app, barra_progreso_rango, archivo_excel_global, hojas_a_convertir
        )
    else:
        messagebox.showerror("Error", "Por favor, seleccione un archivo Excel primero.")


# -------------------------------------------------------------------------------

# Creación de la ventana principal
app = tk.Tk()
app.title("Convertir Excel a PDF")

window_width = 300
window_height = 350
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
center_x = int(screen_width / 2 - window_width / 2)
center_y = int(screen_height / 2 - window_height / 2)
app.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

# -------------------------------------------------------------------------------

pestanas = ttk.Notebook(app)

pestaña_archivos = ttk.Frame(pestanas)
pestaña_rango = ttk.Frame(pestanas)
pestaña_rev = ttk.Frame(pestanas)
ventana_info = ttk.Frame(pestanas)

pestanas.add(pestaña_archivos, text="Documento")
pestanas.add(pestaña_rango, text="Rango")
pestanas.add(pestaña_rev, text="Revisión")
pestanas.add(ventana_info, text="Informacion")

pestanas.pack(expand=True, fill="both")

# -------------------------------------------------------------------------------

# Pestaña Archivos
convertir_button = ttk.Button(
    pestaña_archivos,
    text="Convertir Excel a PDF",
    command=lambda: convertir_excel_a_pdf(app, barra_progreso_archivos),
)
convertir_button.pack(pady=5)
barra_progreso_archivos = BarraDeCarga(pestaña_archivos)
barra_progreso_archivos.mostrar()
boton_abrir_carpeta_archivos = ttk.Button(
    pestaña_archivos, text="Abrir Carpeta PDF", command=abrir_carpeta_pdf
)
boton_abrir_carpeta_archivos.pack(pady=5)
# -------------------------------------------------------------------------------
# Pestaña Rango
boton_abrir_excel = ttk.Button(
    pestaña_rango, text="Abrir Archivo Excel", command=abrir_archivo_excel
)
boton_abrir_excel.pack(pady=5)
label_hoja_inicio = ttk.Label(pestaña_rango, text="Hoja Inicio:")
label_hoja_inicio.pack(pady=5)
lista_hojas_inicio = ttk.Combobox(pestaña_rango)
lista_hojas_inicio.pack(pady=5)
label_hoja_fin = ttk.Label(pestaña_rango, text="Hoja Fin:")
label_hoja_fin.pack(pady=5)
lista_hojas_fin = ttk.Combobox(pestaña_rango)
lista_hojas_fin.pack(pady=5)
boton_convertir_rango = ttk.Button(
    pestaña_rango,
    text="Convertir",
    command=lambda: convertir_rango_especifico(
        app, lista_hojas_inicio.get(), lista_hojas_fin.get(), barra_progreso_rango
    ),
)
boton_convertir_rango.pack(pady=5)
barra_progreso_rango = BarraDeCarga(pestaña_rango)
barra_progreso_rango.mostrar()
boton_abrir_carpeta_rango = ttk.Button(
    pestaña_rango, text="Abrir Carpeta PDF", command=abrir_carpeta_pdf
)
boton_abrir_carpeta_rango.pack(pady=5)
# -------------------------------------------------------------------------------
# Pestaña revision
boton_borrar_archivos_repetidos = ttk.Button(
    pestaña_rev,
    text="Borrar Archivos Repetidos",
    command=reemplazar_espacios_en_archivos,
)
boton_borrar_archivos_repetidos.pack(
    pady=5
)  # Asegúrate de usar .pack() o .grid() para mostrar el botón
# -------------------------------------------------------------------------------

# Pestaña Informacion
boton_abrir_excel = ttk.Button(
    ventana_info, text="Abrir Archivo Excel", command=abrir_archivo_excel
)
label_hoja_inicio = ttk.Label(ventana_info, text="Nombre: Giovanny prens luna")
label_hoja_inicio.pack(pady=18)


app.mainloop()
